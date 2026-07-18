import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment,Border, Side, Font, PatternFill
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.pagebreak import Break
from datetime import datetime
from flask import Flask, render_template, redirect, url_for ,request,send_file
import sqlite3 
from dateutil.relativedelta import relativedelta
import pandas as pd 

app = Flask(__name__)   

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATABASE_PATH = os.path.join(BASE_DIR, "database", "students.db")

def create_table():
    conn = sqlite3.connect(DATABASE_PATH)
    conn.execute("PRAGMA journal_mode=WAL;")
    cursor = conn.cursor()
    cursor.execute(""" CREATE TABLE IF NOT EXISTS students
                ( id INTEGER PRIMARY KEY AUTOINCREMENT,
                စဉ် TEXT,
                ကျောင်းဝင်အမှတ် TEXT,
                နာမည် TEXT,
                အဖေနာမည် TEXT,
                ကျားမ TEXT,
                မွေးနေ့ TEXT,
                class TEXT,
                UNIQUE(ကျောင်းဝင်အမှတ်) ) """)
    
    conn.commit()
    conn.close()

create_table()

def calculate_age_avg(dob_str, date):
    dob = datetime.strptime(dob_str, "%d-%m-%Y")
    target_date = datetime.strptime(date, "%Y-%m-%d")
    check_date = target_date - relativedelta(months=6)
    age = check_date.year - dob.year

    if (check_date.month, check_date.day) < (dob.month, dob.day):
        age -= 1

    return age + 1

def calculate_age(dob_str,date):
     dob = datetime.strptime(dob_str, "%d-%m-%Y") 
     target_date = datetime.strptime(date, "%Y-%m-%d") 
     age = target_date.year - dob.year - ((target_date.month, target_date.day) < (dob.month, dob.day)) 
     
     return age

def apply_filters(data_age, class_html, gender_html, age_html, age_html_1, age_html_2):
    data = data_age
    if class_html != 'all':
        data = [row for row in data if row[6] == class_html]
    if gender_html != 'all':
        data = [row for row in data if row[3] == gender_html]
    if age_html:
        data = [row for row in data if int(row[7]) == int(age_html)]
    elif age_html_1 and age_html_2:
        data = [row for row in data if int(age_html_1) <= int(row[7]) <= int(age_html_2)]
    return data

def eng_to_mm(number):
    eng_digits = ['0','1','2','3','4','5','6','7','8','9']
    mm_digits = ['၀','၁','၂','၃','၄','၅','၆','၇','၈','၉']
    number_list = []
    for num in number:
        for eng,mm in zip(eng_digits,mm_digits):
            if num == mm:
                number_list.append(num)
            elif num == eng:
                number_list.append(mm)

    return ''.join(number_list)

@app.route('/')
def home():
    return render_template('try.html')

@app.route('/insert', methods=['POST'])
def insert_data():
    excel_file = request.files['myfile']
    with sqlite3.connect(DATABASE_PATH, timeout=15) as conn:
        cursor = conn.cursor()
    
        cursor.execute("DELETE FROM students")
    
        df = pd.read_excel(excel_file,dtype={'စဉ်':str,'ကျောင်းဝင်အမှတ်':str})
        df['စဉ်'] = df['စဉ်'].apply(eng_to_mm)
        df['ကျောင်းဝင်အမှတ်'] = df['ကျောင်းဝင်အမှတ်'].apply(eng_to_mm)
        df.columns = df.columns.str.strip()
    
        for _, row in df.iterrows():
            cursor.execute(""" INSERT OR IGNORE INTO students 
            (စဉ်,ကျောင်းဝင်အမှတ်,နာမည်,ကျားမ,အဖေနာမည်,မွေးနေ့,class) 
            VALUES (?, ?, ?, ?, ?, ?, ?) """,
           (row['စဉ်'],
            row['ကျောင်းဝင်အမှတ်'],
            row['နာမည်'],
            row['ကျားမ'],
            row['အဖေနာမည်'],
            pd.to_datetime(row['မွေးနေ့']).strftime('%d-%m-%Y'),
            row['class']))
    return redirect(url_for('home'))

@app.route('/view_tb', methods=['POST'])
def view_data():
    age_html = request.form.get('age')
    age_html_1 = request.form.get('age_1')
    age_html_2 = request.form.get('age_2')
    class_html = request.form.get('class')
    gender_html = request.form.get('gender')
    filename   = request.form.get('file_name')
    file_type = request.form.get('file_type')
    school_name_html = request.form.get('school_name')
    action_html = request.form.get('on')
    date = request.form.get('date')

    if action_html == 'run':
        if age_html == '' and age_html_1 == '' and age_html_2 == '':
            conn = sqlite3.connect(DATABASE_PATH, timeout=10)
            cursor = conn.cursor()
            cursor.execute("SELECT စဉ်, ကျောင်းဝင်အမှတ်, နာမည်, ကျားမ, အဖေနာမည်, မွေးနေ့ ,class FROM students")
            rows = cursor.fetchall()
            conn.close()
            
            data_age = []
            for row in rows:
                no, roll, name, gender, father, dob, class_name = row
                age = calculate_age_avg(dob, date)
                data_age.append((no, roll, name, gender, father, dob, class_name, age))

            data = apply_filters(data_age, class_html, gender_html, age_html, age_html_1, age_html_2)
        elif age_html != '':
            conn = sqlite3.connect(DATABASE_PATH, timeout=10)
            cursor = conn.cursor()
            cursor.execute("SELECT စဉ်, ကျောင်းဝင်အမှတ်, နာမည်, ကျားမ, အဖေနာမည်, မွေးနေ့ ,class FROM students")
            rows = cursor.fetchall()
            conn.close()
            
            data_age = []
            for row in rows:
                no, roll, name, gender, father, dob, class_name = row
                age = calculate_age_avg(dob, date)
                data_age.append((no, roll, name, gender, father, dob, class_name, age))

            data = apply_filters(data_age, class_html, gender_html, age_html, age_html_1, age_html_2)

        elif age_html_1 != '' or age_html_2 != '':
            conn = sqlite3.connect(DATABASE_PATH, timeout=10)
            cursor = conn.cursor()
            cursor.execute("SELECT စဉ်, ကျောင်းဝင်အမှတ်, နာမည်, ကျားမ, အဖေနာမည်, မွေးနေ့ ,class FROM students")
            rows = cursor.fetchall()
            conn.close()
            
            data_age = []
            for row in rows:
                no, roll, name, gender, father, dob, class_name = row
                age = calculate_age(dob, date)
                data_age.append((no, roll, name, gender, father, dob, class_name, age))

            data = apply_filters(data_age, class_html, gender_html, age_html, age_html_1, age_html_2)

        return render_template('try.html', data=data,
                               data_male=len([r for r in data if r[3] == 'ကျား']),
                               data_female=len([r for r in data if r[3] == 'မ']),
                               data_all=len(data),
                               age=age_html, age_1=age_html_1, age_2=age_html_2,
                               class_html=class_html, gender=gender_html, date=date)
    
    elif action_html == 'excel_download':
        if file_type == 'normal':
            download_data = [
                {
                    'စဉ်': r[0],
                    'ကျောင်းဝင်အမှတ်': r[1],
                    'နာမည်': r[2],
                    'ကျားမ': r[3],
                    'အဖေနာမည်': r[4],
                    'မွေးနေ့': r[5],
                    'class': r[6],
                    'အသက်': r[7]
                }for r in data 
            ]
            df = pd.DataFrame(download_data)
            excel_filename = filename + '.xlsx'
            df.to_excel(excel_filename, index=False)
            wb = load_workbook(excel_filename)
            ws = wb.active
            ws.column_dimensions['A'].width = 6  #စဉ်
            ws.column_dimensions['B'].width = 18 #ကျောင်းဝင်အမှတ်
            ws.column_dimensions['C'].width = 26 #အမည်
            ws.column_dimensions['D'].width = 8 #ကျား/မ
            ws.column_dimensions['E'].width = 26  #အ‌ဖေနာမည်
            ws.column_dimensions['F'].width = 12 #မွေးနေ့
            ws.column_dimensions['G'].width = 8 #class
            ws.column_dimensions['H'].width = 8 #အသက်

            for row in ws.iter_rows():
                ws.row_dimensions[row[0].row].height = 20
            
            for cell in ws[1]:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            for row in ws.iter_rows(min_row=2):
                row[0].alignment = Alignment(horizontal='center', vertical='center') #စဉ်
                row[1].alignment = Alignment(horizontal='center', vertical='center') #ကျောင်းဝင်အမှတ်
                row[2].alignment = Alignment(horizontal='left', vertical='center')  #နာမည်
                row[3].alignment = Alignment(horizontal='center', vertical='center') #ကျားမ
                row[4].alignment = Alignment(horizontal='left', vertical='center')  #အဖေနာမည်
                row[5].alignment = Alignment(horizontal='center', vertical='center') #မွေးနေ့
                row[6].alignment = Alignment(horizontal='center', vertical='center') #class
                row[7].alignment = Alignment(horizontal='center', vertical='center') #အသက်

            wb.save(excel_filename)

            return send_file(excel_filename, as_attachment=True)
        elif file_type == 'normalized':
            no, school_name, roll, name, gender, father, dob, reasion_ = [], [], [], [], [], [], [], []
            for x in data:
                no.append(x[0])
                school_name.append(school_name_html)
                roll.append(x[1])
                name.append(x[2])
                gender.append(x[3])
                father.append(x[4])
                dob.append(x[5])
                reasion_.append('')

            download_data = {
                'စဉ်': no,
                'ကျောင်းအမည်': school_name,
                'ကျောင်းဝင်အမှတ်': roll,
                'အမည်': name,
                'ကျား/မ': gender,
                'အဖေအမည်': father,
                'မွေးသက္ကရာဇ်': dob,
                'မှတ်ချက်': reasion_
            }
            df = pd.DataFrame(download_data)
            excel_filename = filename + '_normalized.xlsx'
            df.to_excel(excel_filename, index=False)
            wb = load_workbook(excel_filename)
            ws = wb.active

            thick = Side(style='thin')
            border = Border(
                left=thick,
                right=thick,
                top=thick,
                bottom=thick
            )

            for row in ws.iter_rows():
                for cell in row:
                    cell.border = border

            bold_font = Font(bold=True)

            for cell in ws[1]:
                cell.font = bold_font

            ws.insert_rows(1, amount=4)

            ws['H1'] = "ပူးတွဲ(က)"
            ws['A2'] = "အခြေခံပညာဦးစီးဌာန"
            ws['A3'] = "ရန်ကုန်တိုင်းဒေသကြီးအင်းစိန်ခရိုင်လှိုင်သာယာ(အနောက်ပိုင်း)မြို့နယ်"
            ws['A4'] = "၂၀၂၆ - ၂၀၂၇ ပညာသင်နှစ်၊ အခြေခံပညာ မူလတန်းအဆင့် Grade5 စာမေးပွဲဖြေဆိုသူစာရင်းပေါင်းချုပ်"

            ws.merge_cells('A2:H2')
            ws.merge_cells('A3:H3')
            ws.merge_cells('A4:H4')

            for row in range(1, 5):
                cell = ws[f'A{row}']

                cell.alignment = Alignment(
                    horizontal='center',
                    vertical='center'
                )

                cell.font = Font(
                    bold=True,
                    size=13
                )

            ws['H1'].font = Font(
                bold=True,
                size=13
            )
            ws['H1'].alignment = Alignment(
                horizontal='center',
                vertical='center'
            )

            header_fill = PatternFill(
                start_color="D9EAD3",
                end_color="D9EAD3",
                fill_type="solid"
            )

            for cell in ws[5]:
                cell.font = Font(bold=True)

                cell.alignment = Alignment(
                    horizontal='center',
                    vertical='center'
                )

                cell.fill = header_fill

            ws.column_dimensions['A'].width = 6  #စဉ်
            ws.column_dimensions['B'].width = 14 #ကျောင်းအမည်
            ws.column_dimensions['C'].width = 18 #ကျောင်းဝင်အမှတ်
            ws.column_dimensions['D'].width = 26 #အမည်
            ws.column_dimensions['E'].width = 8  #ကျား/မ
            ws.column_dimensions['F'].width = 26 #အဖေအမည်
            ws.column_dimensions['G'].width = 15 #မွေးနေ့
            ws.column_dimensions['H'].width = 22 #မှတ်ချက်

            for row in ws.iter_rows():
                ws.row_dimensions[row[0].row].height = 23

            for row in ws.iter_rows(min_row=5):
                ws.row_dimensions[row[0].row].height = 19.4

            for row in ws.iter_rows():
                for cell in row:
                    cell.font = Font(
                        name='Pyidaungsu',
                        size=13,
                        bold=True
                    )
            for row in ws.iter_rows(min_row=5):
                for cell in row:
                    cell.font = Font(
                        name='Pyidaungsu',
                        size=11
                    )            

            for row in ws.iter_rows(min_row=6):

                row[0].alignment = Alignment(horizontal='center', vertical='center')
                row[1].alignment = Alignment(horizontal='left', vertical='center')
                row[2].alignment = Alignment(horizontal='center', vertical='center')
                row[3].alignment = Alignment(horizontal='left', vertical='center')
                row[4].alignment = Alignment(horizontal='center', vertical='center')
                row[5].alignment = Alignment(horizontal='left', vertical='center')
                row[6].alignment = Alignment(horizontal='center', vertical='center')
                row[7].alignment = Alignment(horizontal='left', vertical='center')

            thin = Side(style='thin')

            border = Border(
                left=thin,
                right=thin,
                top=thin,
                bottom=thin
            )

            for row in ws.iter_rows(min_row=5):
                for cell in row:
                    cell.border = border

            ws.freeze_panes = 'A6'

            ws.page_setup.paperSize = ws.PAPERSIZE_A4
            ws.page_setup.orientation = 'portrait'

            ws.page_margins = PageMargins(
                left=0.3,
                right=0.3,
                top=0.5,
                bottom=0.5
            )

            ws.print_title_rows = '1:5'

            ws.print_options.horizontalCentered = True

            data_start_row = 5      # data starts at row 6
            rows_per_page = 22

            current_row = data_start_row + rows_per_page

            while current_row <= ws.max_row:
                ws.row_breaks.append(Break(id=current_row))
                current_row += rows_per_page

            wb.save(excel_filename)
            return send_file(excel_filename, as_attachment=True)

    return render_template('try.html')

if __name__ == '__main__':
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 5000)))
